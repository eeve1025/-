import openpyxl
#抓取"104有關大數據工作"相關資料--作品1
import requests
import bs4

res=requests.get('https://www.104.com.tw/jobs/search/?ro=0&keyword=%E5%A4%A7%E6%95%B8%E6%93%9A&expansionType=area%2Cspec%2Ccom%2Cjob%2Cwf%2Cwktm&order=1&asc=0&page=1&mode=s&jobsource=2018indexpoc&langFlag=0&langStatus=0&recommendJob=1&hotJob=1')
soup=bs4.BeautifulSoup(res.text)

wb=openpyxl.Workbook()
ws=wb.active

ws['A1']='職缺名稱'
ws['B1']='職缺連結'
ws['C1']='公司名稱'
ws['D1']='工作地區'
ws['E1']='薪資待遇'

page=1
while soup.find_all('article',class_="b-block--top-bord job-list-item b-clearfix js-job-item")!=[]:
  print("=======================")
  print("現在正在讀取第",page,"頁")
  print("=======================")

  for job in soup.find_all('article',class_="b-block--top-bord job-list-item b-clearfix js-job-item"):

    a=job.a.text
    b='http:'+job.a['href']
    c=job.ul.a.text.strip()
    d=job.select('ul')[1].li.text
    if  job.find('div',class_="job-list-tag b-content").select('span')!=[] and job.find('div',class_="job-list-tag b-content").span.text=="待遇面議":
        e=job.find('div',class_="job-list-tag b-content").span.text
    else:
        e=job.find('div',class_="job-list-tag b-content").a.text

        f=e[:2]
        salary=''
        for char in e:
          if char.isdigit() or char=='~':
            salary+=char

        #薪水有範圍值，拆成上下限值
        if '~' in salary:
            #salary1=e[4]
          low_price=salary[:salary.find('~')]
          high_price=salary[salary.find('~')+1:]
        else:
          low_price=salary
          high_price=salary 

        #轉成數字格式
        low_price=int(low_price)
        high_price=int(high_price)

    ws.append([a,b,c,d,e,f,low_price,high_price])
  page+=1
  res=requests.get('https://www.104.com.tw/jobs/search/?ro=0&kwop=7&keyword=%E5%A4%A7%E6%95%B8%E6%93%9A&expansionType=area%2Cspec%2Ccom%2Cjob%2Cwf%2Cwktm&order=12&asc=0&page='+str(page)+'&mode=s&jobsource=2018indexpoc&langFlag=0&langStatus=0&recommendJob=1&hotJob=1')
  soup=bs4.BeautifulSoup(res.text)
  wb.save("104大數據職缺清單0812.xlsx")



